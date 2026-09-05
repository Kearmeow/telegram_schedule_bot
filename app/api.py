from io import BytesIO
from typing import Any

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from . import db
from .config import ADMIN_IDS

app = FastAPI(title="Telegram Schedule Mini App")


class GroupCreate(BaseModel):
    name: str = Field(min_length=1, max_length=100)


class GroupSelect(BaseModel):
    telegram_id: int
    group_id: int


class LessonCreate(BaseModel):
    group_id: int
    weekday: int = Field(ge=1, le=7)
    start_time: str = Field(min_length=1, max_length=10)
    end_time: str = Field(min_length=1, max_length=10)
    subject: str = Field(min_length=1, max_length=200)
    teacher: str = Field(default="", max_length=200)
    room: str = Field(default="", max_length=100)
    notes: str = Field(default="", max_length=500)


class LessonUpdate(LessonCreate):
    pass

class TextScheduleImport(BaseModel):
    group: str = Field(min_length=1, max_length=100)
    text: str = Field(min_length=1, max_length=30000)
    mode: str = Field(default="replace", pattern="^(append|replace)$")


def is_admin(telegram_id: int) -> bool:
    return telegram_id in ADMIN_IDS


def require_admin(telegram_id: int):
    if not is_admin(telegram_id):
        raise HTTPException(status_code=403, detail="Нет доступа к админке")


DAY_MAP = {
    "1": 1, "пн": 1, "понедельник": 1,
    "2": 2, "вт": 2, "вторник": 2,
    "3": 3, "ср": 3, "среда": 3,
    "4": 4, "чт": 4, "четверг": 4,
    "5": 5, "пт": 5, "пятница": 5,
    "6": 6, "сб": 6, "суббота": 6,
    "7": 7, "вс": 7, "воскресенье": 7,
}

HEADER_ALIASES = {
    "group": {"group", "группа", "группы"},
    "weekday": {"day", "weekday", "день", "день недели"},
    "start_time": {"start", "start_time", "начало", "время начала", "начало пары"},
    "end_time": {"end", "end_time", "конец", "время конца", "конец пары"},
    "subject": {"subject", "предмет", "дисциплина"},
    "teacher": {"teacher", "преподаватель", "препод"},
    "room": {"room", "аудитория", "кабинет", "каб"},
    "notes": {"notes", "note", "заметки", "заметка", "примечание"},
}


def normalize(value: Any) -> str:
    return "" if value is None else str(value).strip()


def parse_day(value: Any) -> int:
    text = normalize(value).lower().replace("ё", "е")
    if text in DAY_MAP:
        return DAY_MAP[text]
    raise ValueError(f"Неизвестный день: {value}")


def parse_time(value: Any) -> str:
    text = normalize(value)
    if not text:
        raise ValueError("Пустое время")
    if hasattr(value, "strftime"):
        return value.strftime("%H:%M")
    # Excel may return values such as 08:30:00 or 8:30 AM.
    if len(text) >= 5 and text[2] == ":" and text[:2].isdigit():
        return text[:5]
    return text


def header_map(headers: list[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    normalized = [normalize(h).lower().replace("ё", "е") for h in headers]
    for key, aliases in HEADER_ALIASES.items():
        for index, header in enumerate(normalized):
            if header in aliases:
                result[key] = index
                break
    required = ["group", "weekday", "start_time", "end_time", "subject"]
    missing = [x for x in required if x not in result]
    if missing:
        raise ValueError("Не найдены обязательные колонки: " + ", ".join(missing))
    return result


@app.on_event("startup")
async def startup():
    db.init_db()


@app.get("/")
async def index():
    return FileResponse("web/index.html")


@app.get("/app.js")
async def javascript():
    return FileResponse("web/app.js", media_type="application/javascript")


@app.get("/style.css")
async def css():
    return FileResponse("web/style.css", media_type="text/css")


@app.get("/api/groups")
async def groups():
    return db.list_groups()


@app.get("/api/groups/{group_id}/schedule")
async def schedule(group_id: int, weekday: int | None = None):
    if not db.get_group(group_id):
        raise HTTPException(404, "Группа не найдена")
    if weekday is not None and not 1 <= weekday <= 7:
        raise HTTPException(400, "Некорректный день")
    return db.get_lessons(group_id, weekday)


@app.post("/api/users/group")
async def select_group(data: GroupSelect):
    if not db.get_group(data.group_id):
        raise HTTPException(404, "Группа не найдена")
    db.set_user_group(data.telegram_id, data.group_id)
    return {"ok": True}


@app.get("/api/users/{telegram_id}/group")
async def user_group(telegram_id: int):
    return {"group_id": db.get_user_group(telegram_id)}


@app.get("/api/admin/check/{telegram_id}")
async def admin_check(telegram_id: int):
    return {"admin": is_admin(telegram_id)}


@app.post("/api/admin/groups")
async def add_group(data: GroupCreate, telegram_id: int):
    require_admin(telegram_id)
    try:
        group_id = db.create_group(data.name)
    except Exception:
        raise HTTPException(400, "Не удалось создать группу: возможно, она уже существует")
    return {"id": group_id, "name": data.name.strip()}


@app.delete("/api/admin/groups/{group_id}")
async def remove_group(group_id: int, telegram_id: int):
    require_admin(telegram_id)
    if not db.get_group(group_id):
        raise HTTPException(404, "Группа не найдена")
    db.delete_group(group_id)
    return {"ok": True}


@app.post("/api/admin/lessons")
async def add_lesson(data: LessonCreate, telegram_id: int):
    require_admin(telegram_id)
    if not db.get_group(data.group_id):
        raise HTTPException(404, "Группа не найдена")
    lesson_id = db.create_lesson(
        data.group_id, data.weekday, data.start_time, data.end_time,
        data.subject.strip(), data.teacher.strip(), data.room.strip(), data.notes.strip()
    )
    return {"id": lesson_id}


@app.put("/api/admin/lessons/{lesson_id}")
async def edit_lesson(lesson_id: int, data: LessonUpdate, telegram_id: int):
    require_admin(telegram_id)
    db.update_lesson(
        lesson_id, data.weekday, data.start_time, data.end_time,
        data.subject.strip(), data.teacher.strip(), data.room.strip(), data.notes.strip()
    )
    return {"ok": True}


@app.delete("/api/admin/lessons/{lesson_id}")
async def remove_lesson(lesson_id: int, telegram_id: int):
    require_admin(telegram_id)
    db.delete_lesson(lesson_id)
    return {"ok": True}



def parse_text_schedule(text: str) -> list[dict[str, Any]]:
    """Parse a compact human-friendly schedule format.

    Supported examples:
      Понедельник
      08:30-10:05 Математика | Иванов И.И. | 301
      10:15-11:50 Информатика | Петров П.П. | 205

    Also supports:
      1 | 08:30-10:05 | Математика | Иванов И.И. | 301

    Empty lines and comments beginning with # are ignored.
    """
    import re

    lines = [line.strip() for line in text.splitlines()]
    current_day = None
    result: list[dict[str, Any]] = []

    time_re = re.compile(r"^(\d{1,2})[.:](\d{2})\s*[-–—]\s*(\d{1,2})[.:](\d{2})(?:\s+|$)")

    for line_no, line in enumerate(lines, start=1):
        if not line or line.startswith("#"):
            continue

        # A standalone day heading.
        try:
            current_day = parse_day(line)
            continue
        except ValueError:
            pass

        parts = [p.strip() for p in line.split("|")]

        # Format: lesson_number | time | subject | teacher | room | notes
        if len(parts) >= 3 and re.fullmatch(r"\d{1,2}", parts[0]):
            if current_day is None:
                raise ValueError(f"Строка {line_no}: сначала укажи день недели")
            time_match = re.fullmatch(
                r"(\d{1,2}[.:]\d{2})\s*[-–—]\s*(\d{1,2}[.:]\d{2})", parts[1]
            )
            if not time_match:
                raise ValueError(f"Строка {line_no}: неверное время «{parts[1]}»")
            h1, m1, h2, m2 = time_match.groups()
            start_time, end_time = f"{int(h1):02d}:{m1}", f"{int(h2):02d}:{m2}"
            subject = parts[2]
            teacher = parts[3] if len(parts) > 3 else ""
            room = parts[4] if len(parts) > 4 else ""
            notes = parts[5] if len(parts) > 5 else ""
        else:
            # Format: time subject | teacher | room | notes
            time_match = time_re.match(line)
            if not time_match:
                raise ValueError(
                    f"Строка {line_no}: не удалось распознать занятие. "
                    "Пример: 08:30-10:05 Математика | Иванов | 301"
                )
            h1, m1, h2, m2 = time_match.groups()
            start_time, end_time = f"{int(h1):02d}:{m1}", f"{int(h2):02d}:{m2}"
            rest = line[time_match.end():].strip()
            fields = [p.strip() for p in rest.split("|")]
            subject = fields[0] if fields else ""
            teacher = fields[1] if len(fields) > 1 else ""
            room = fields[2] if len(fields) > 2 else ""
            notes = fields[3] if len(fields) > 3 else ""

        if not subject:
            raise ValueError(f"Строка {line_no}: не указан предмет")

        result.append({
            "weekday": current_day,
            "start_time": start_time,
            "end_time": end_time,
            "subject": subject,
            "teacher": teacher,
            "room": room,
            "notes": notes,
        })

    if not result:
        raise ValueError("Не найдено ни одного занятия")

    return result


@app.post("/api/admin/import/text")
async def import_text_schedule(data: TextScheduleImport, telegram_id: int):
    require_admin(telegram_id)

    group_name = data.group.strip()
    parsed = parse_text_schedule(data.text)

    # Reuse the DB import mechanism. It expects a "group" field in every row.
    rows = [{**lesson, "group": group_name} for lesson in parsed]
    try:
        result = db.import_lessons(rows, replace=data.mode == "replace")
        return {"ok": True, **result}
    except Exception as exc:
        raise HTTPException(400, f"Не удалось сохранить расписание: {exc}")


@app.get("/api/admin/import/template")
async def import_template(telegram_id: int):
    require_admin(telegram_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    ws.append(["Группа", "День", "Начало", "Конец", "Предмет", "Преподаватель", "Аудитория", "Заметки"])
    ws.append(["ИС-21", "Пн", "08:30", "10:00", "Математика", "Иванов И.И.", "301", ""])
    ws.append(["ИС-21", "Пн", "10:10", "11:40", "Информатика", "Петров П.П.", "205", ""])
    ws.freeze_panes = "A2"
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="schedule_template.xlsx"'},
    )


@app.post("/api/admin/import")
async def import_excel(
    telegram_id: int,
    file: UploadFile = File(...),
    mode: str = Query("append", pattern="^(append|replace)$"),
):
    require_admin(telegram_id)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Нужен файл Excel в формате .xlsx")

    try:
        raw = await file.read()
        workbook = load_workbook(BytesIO(raw), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        if not rows:
            raise ValueError("Файл пустой")

        mapping = header_map(list(rows[0]))
        parsed: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows[1:], start=2):
            if not any(normalize(v) for v in row):
                continue
            group = normalize(row[mapping["group"]])
            subject = normalize(row[mapping["subject"]])
            if not group or not subject:
                raise ValueError(f"Строка {row_number}: группа и предмет обязательны")
            parsed.append({
                "group": group,
                "weekday": parse_day(row[mapping["weekday"]]),
                "start_time": parse_time(row[mapping["start_time"]]),
                "end_time": parse_time(row[mapping["end_time"]]),
                "subject": subject,
                "teacher": normalize(row[mapping["teacher"]]) if "teacher" in mapping else "",
                "room": normalize(row[mapping["room"]]) if "room" in mapping else "",
                "notes": normalize(row[mapping["notes"]]) if "notes" in mapping else "",
            })

        if not parsed:
            raise ValueError("В Excel нет занятий")

        result = db.import_lessons(parsed, replace=mode == "replace")
        return {"ok": True, **result}
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Не удалось импортировать Excel: {exc}")
